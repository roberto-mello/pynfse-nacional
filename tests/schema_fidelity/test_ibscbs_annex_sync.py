"""ANEXO_C sync for IBSCBS operation tables."""

from __future__ import annotations

from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

from pynfse_nacional.models_ibscbs import (
    IBSCBS_C_IND_OP_ALLOWLIST,
    IBSCBS_C_IND_OP_CODES,
    IBSCBS_OPERATION_CATEGORIES,
)

FIXTURE_PATH = (
    Path(__file__).resolve().parents[1]
    / "fixtures"
    / "annexes"
    / "anexo_c-indop_ibscbs-snnfse-v1-01-20260122.xlsx"
)


def _normalize_text(value: object | None) -> str:
    if value is None:
        return ""

    return " ".join(str(value).split())


def _normalize_layout(value: object | None) -> str:
    if value is None:
        return ""

    lines = [line.strip() for line in str(value).splitlines() if line.strip()]
    paths = [
        line
        for line in lines
        if line.startswith("NFSe/") or line.startswith("EXCLUSIVO")
    ]

    if paths:
        return " | ".join(paths)

    return " ".join(lines)


def _parse_annex_c() -> tuple[dict[str, tuple[object, ...]], tuple[str, ...], set[str]]:
    workbook = load_workbook(FIXTURE_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    categories: "OrderedDict[str, list[tuple[str, str, str, str]]]" = OrderedDict()
    variants: list[str] = []
    allowlist: set[str] = set()
    current_inciso = ""
    current_codigo_base = ""

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        c_ind_op = row[6]
        if c_ind_op is None:
            continue

        inciso_cell = _normalize_text(row[0])
        if inciso_cell:
            current_inciso = inciso_cell.removeprefix("Inc. ").strip()

        code = str(c_ind_op).zfill(6)
        codigo_base_cell = _normalize_text(row[4])
        if codigo_base_cell:
            if code == "080101":
                current_codigo_base = "0801"
            else:
                current_codigo_base = codigo_base_cell.zfill(4)

        sequence = code[4:]
        variants.append(code)

        if code != "080101":
            allowlist.add(code)

        codigo_base = current_codigo_base

        category_rows = categories.setdefault(codigo_base, [])
        category_rows.append(
            (
                current_inciso,
                _normalize_text(row[1]),
                _normalize_text(row[2]),
                _normalize_text(row[3]),
                codigo_base,
                sequence,
                code,
                _normalize_text(row[7]),
                _normalize_layout(row[8]),
            )
        )

    category_signatures: dict[str, tuple[object, ...]] = {}
    for codigo_base, rows in categories.items():
        first = rows[0]
        category_signatures[codigo_base] = (
                "Art. 11",
                first[0],
                first[1],
                first[2],
                first[3],
                codigo_base,
                tuple(
                    (
                        row[5],
                        row[6],
                        row[7],
                        row[8],
                    )
                    for row in rows
                ),
        )

    return category_signatures, tuple(variants), allowlist


def _category_signature(category) -> tuple[object, ...]:
    return (
        _normalize_text(category.artigo),
        _normalize_text(category.inciso),
        _normalize_text(category.tipo_operacao),
        _normalize_text(category.local_operacao),
        _normalize_text(category.caracteristica_fornecimento),
        category.codigo_base,
        tuple(
            (
                variant.sequencia,
                variant.c_ind_op,
                _normalize_text(variant.local_fornecimento),
                _normalize_text(variant.campo_layout),
            )
            for variant in category.variantes
        ),
    )


def test_annex_c_matches_ibscbs_operation_tables():
    expected_categories, expected_codes, expected_allowlist = _parse_annex_c()

    actual_categories = {
        category.codigo_base: _category_signature(category)
        for category in IBSCBS_OPERATION_CATEGORIES
    }

    assert actual_categories == expected_categories
    assert set(IBSCBS_C_IND_OP_CODES) == set(expected_codes)
    assert IBSCBS_C_IND_OP_ALLOWLIST == expected_allowlist
